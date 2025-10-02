# app.py
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


st.markdown(
    """
    <h1 style="text-align:center; color:#d35400;">🍕 Transformateur de commandes</h1>
    <p style="font-size:16px;">
        Petit guide sur comment vous devez faire.<br>
        <b> C'est pas compliqué :</b>
    </p>
    <ol>
        <li> Mettez le fichier que vous recevez de Rico ( un truc du genre ordrerxxxx (.xlsx) </li>
        <li>Télécharger le fichier transformé qui va s'appeler un truc du style commande traitées</li><br></br>
    <p> Ensuite, vous ouvrez le super document WORD du drive  </p>
        <li> A l'ouverture du fichier il va probablement vous demander un document à mettre. Télécharger le document "Tab à pizza" du drive et mettez le quand Word vous le demande</li>
        <li> Ensuite quand le document Word est bien ouvert ( si il vous demande des trucs appuyer sur ok à chaque fois), appuyez sur l'onglet publipostage</li>
        <li> Dans l'onglet publipostage, il faut appuyer sur - Sélectionner des destinataires -> Utiliser une liste existante </li>
        <li> Ensuite, choisissez bien cette fois le fichier transformé ("Commandes_traitées")
        <li> Si tout fonctionne, vous pouvez maintenant parcourir les différentes commandes des gens en appuyant sur "Apercu des résultats" et les flèches </li>
        <li> Il vous reste que à exporter le résultat à l'aide de "Terminer et fusionnez tout à droite"</li><br></br>   

    <p>Normalement tout est bon à ce stade, je vous conseille de vérifier quand même 2-3 commandes si tout correspond bien !<br>
        Si jamais y'a un bad, dites moi, je devrais pouvoir corriger ça rapidement. Et si jamais, j'ai déjà mis à jour le doc Word par rapport aux nouvelles sortes de pizzas donc c'est bon normalement ! </p>
    </ol>
    """,
    unsafe_allow_html=True
)

# Étape 1 : Upload du fichier Excel
fichier = st.file_uploader("Importe le fichier Excel brut (.xlsx)", type=["xlsx"])

if fichier is not None:
    # Charger le fichier dans un DataFrame
    df = pd.read_excel(fichier)

    # ⚡ Exemple de traitement (à remplacer par ta vraie fonction)
    def heure(row):
        if "17-18" in row["SKU"]:
            return "18h - 18h30"
        if "18-19" in row["SKU"]:
            return "18h30 - 19h30"
        if "19-20" in row["SKU"]:
            return "19h30 - 20h30"
        if "20-21" in row["SKU"]:
            return "20h30 - 21h30"

    df['No de commande'] = df['Order Number']
    df['Nom'] = df['First Name (Billing)']
    df['Prénom'] = df["Last Name (Billing)"]
    df['Adresse'] = df['Address 1&2 (Billing)']
    df['Téléphone'] = df['Phone (Billing)']
    df['Heure de livraison'] = df.apply(heure, axis=1)
    df["Nom item"]=df["Item Name"]
    df["remarque"]=df["Customer Note"]

    def hawai (row):
        if "hawa" in row["Nom item"].lower() :
            return row["Quantity"]
    def prosciuto (row):
        if ("pr") in row["Nom item"].lower() :
            if not "fu" in row["Nom item"].lower():
                return row["Quantity"]
    def funghi (row):
        if ("fun") in row["Nom item"].lower() :
            if not "pro" in row["Nom item"].lower():
                return row["Quantity"]
    def prosciuto_ef (row):
        if ("pr") in row["Nom item"].lower() :
            if ("fu") in row["Nom item"].lower():
                return row["Quantity"]

    def margherita (row):
        if "marg" in row["Nom item"].lower() :
            return row["Quantity"]

    

    df['hawai'] = df.apply(hawai, axis=1).fillna(0)
    df["jambon"] = df.apply(prosciuto, axis=1).fillna(0)
    df["margherita"] = df.apply(margherita, axis =1).fillna(0)
    df["jambon_champi"] = df.apply(prosciuto_ef, axis=1).fillna(0)
    df["funghi"] = df.apply(funghi, axis=1).fillna(0)
    df["prix_hawai_14"] = df["hawai"] * 14
    df["prix_jambon_14"] = df["jambon"] * 14
    df["prix_margherita_12"] = df["margherita"] * 12
    df["prix_jambon_champi_14"] = df["jambon_champi"] *12
    df["prix_funghi_14"] = df["funghi"]*14
    df["prix total"] = df["prix_hawai_14"] + df["prix_jambon_14"] + df["prix_margherita_12"] + df["prix_jambon_champi_14"] + df["prix_funghi_14"]



    df_new = (
        df.groupby("No de commande")
        .agg({
            "Nom": "first",
            "Prénom": "first",
            "Adresse": "first", 
            "Téléphone": "first",            # même valeur partout → on garde la 1ère
            "Heure de livraison": "first",       # idem
            "hawai": "sum",
            "prix_hawai_14" : "sum",
            "jambon": "sum",
            "prix_jambon_14" : "sum",        # additionne les quantités
            "margherita": "sum",
            "prix_margherita_12" : "sum",
            "jambon_champi": "sum",
            "prix_jambon_champi_14": "sum",
            "funghi": "sum",
            "prix_funghi_14": "sum", 
            "prix total": "sum",
            "remarque": "sum",
        })
        .reset_index()
    )
    zerop=["remarque"]
    df_new[zerop] = df_new[zerop].replace(0, "")
    st.write("V1")
    st.dataframe(df_new.head())
    df_new.to_excel("v1.xlsx", index=False)
   


    from openpyxl import load_workbook
    wb = load_workbook("v1.xlsx")
    ws = wb.active



# Adapter la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # lettre de la colonne
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2  # +2 pour un peu d'espace
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):  # on saute l'en-tête
        if i % 2 == 0:
            for cell in row :
                cell.fill = header_fill


    wb.save("commandes_finales.xlsx")
    # Étape 2 : Afficher un aperçu
    st.write("Aperçu du fichier final transformé 👇")
    st.dataframe(df_new.head())

  

    # Étape 3 : Permettre le téléchargement
    with open("commandes_finales.xlsx", "rb") as f:
        file_data = f.read()

    st.download_button(
        label="⬇️ Télécharger le fichier transformé",
        data=file_data,
        file_name="commandes_traitees.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
